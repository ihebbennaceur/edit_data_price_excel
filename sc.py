import os
import openpyxl
import json
from dotenv import load_dotenv
# import os

import json

# Load the JSON data from the file
with open('category_0.0.3.json', 'r', encoding='utf-8') as file:
    data = json.load(file)

# Iterate through each category and add the new key-value pairs
for category in data['categories']:
    if 'profit' in category:
        category['profit']['UZ'] = "25"
    if 'tax' in category:
        category['tax']['UZ'] = "20"

# Save the updated JSON data back to the file
with open('category_0.0.3.json', 'w', encoding='utf-8') as file:
    json.dump(data, file, ensure_ascii=False, indent=4)

print("Updated JSON file successfully.")


load_dotenv()
products_file = os.getenv('PRODUCTS_FILE')
categories_file = os.getenv('CATEGORIES_FILE')
excel_file = os.getenv('EXCEL_FILE')

def write_to_excel(not_found_data, excel_file):
    existing_fcid = set()
    existing_ncid = set()
    existing_dcid = set()
    existing_frootCatId = set()
    existing_nrootCatId = set()
    existing_drootCatId = set()

    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6, values_only=True):
            if row[0]:  # fcid
                existing_fcid.add(row[0])
            if row[1]:  # ncid
                existing_ncid.add(row[1])
            if row[2]:  # dcid
                existing_dcid.add(row[2])
            if row[3]:  # frootCatId
                existing_frootCatId.add(row[3])
            if row[4]:  # nrootCatId
                existing_nrootCatId.add(row[4])
            if row[5]:  # drootCatId
                existing_drootCatId.add(row[5])
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Not Found Data"
        headers = ["fcid", "ncid", "dcid", "frootCatId", "nrootCatId", "drootCatId"]
        ws.append(headers)

    for i in range(max(len(not_found_data["fcid"]), len(not_found_data["ncid"]), len(not_found_data["dcid"]))):
        fcid = not_found_data["fcid"][i] if i < len(not_found_data["fcid"]) else ""
        ncid = not_found_data["ncid"][i] if i < len(not_found_data["ncid"]) else ""
        dcid = not_found_data["dcid"][i] if i < len(not_found_data["dcid"]) else ""
        frootCatId = not_found_data["frootCatId"][i] if i < len(not_found_data["frootCatId"]) else ""
        nrootCatId = not_found_data["nrootCatId"][i] if i < len(not_found_data["nrootCatId"]) else ""
        drootCatId = not_found_data["drootCatId"][i] if i < len(not_found_data["drootCatId"]) else ""

        
        if fcid and fcid not in existing_fcid:
            ws.append([fcid, "", "", frootCatId, "", ""])
            existing_fcid.add(fcid)
            if frootCatId:
                existing_frootCatId.add(frootCatId)
        if ncid and ncid not in existing_ncid:
            ws.append(["", ncid, "", "", nrootCatId, ""])
            existing_ncid.add(ncid)
            if nrootCatId:
                existing_nrootCatId.add(nrootCatId)
        if dcid and dcid not in existing_dcid:
            ws.append(["", "", dcid, "", "", drootCatId])
            existing_dcid.add(dcid)
            if drootCatId:
                existing_drootCatId.add(drootCatId)

    wb.save(excel_file)


# Function to update product categories
def update_product_categories(products_file, categories_file, excel_file):
    with open(products_file, 'r', encoding='utf-8') as f:
        products = json.load(f)

    with open(categories_file, 'r', encoding='utf-8') as f:
        categories = json.load(f)

    not_found_data = {
        "fcid": [],
        "ncid": [],
        "dcid": [],
        "frootCatId": [],
        "nrootCatId": [],
        "drootCatId": []
    }

    for product in products:
        product_id = product.get('productId', '')
        if not product_id:
            continue

        first_letter = product_id[0]
        if first_letter not in ['f', 'n', 'd']:
            continue

        cid = product.get('cid', None)
        root_cat_id = product.get('rootCatId', None)
        if not cid:
            continue

        category_names = get_category_names(cid, first_letter, categories)
        if not category_names and root_cat_id:
            category_names = get_category_names(root_cat_id, first_letter, categories)

        if category_names:
            product['categoryName'] = {'ru': '|'.join(category_names)}
        else:
            product['categoryName'] = {'ru': ''}
            if first_letter == 'f':
                not_found_data["fcid"].append(cid)
                not_found_data["frootCatId"].append(root_cat_id)
            elif first_letter == 'n':
                not_found_data["ncid"].append(cid)
                not_found_data["nrootCatId"].append(root_cat_id)
            elif first_letter == 'd':
                not_found_data["dcid"].append(cid)
                not_found_data["drootCatId"].append(root_cat_id)

    write_to_excel(not_found_data, excel_file)
    update_size_and_weight(products, categories)

    with open(products_file, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=4)

def get_category_names(cat_id, first_letter, categories):
    category_names = []
    cat_id_str = str(cat_id).rstrip('.0')
    for category in categories.get('categories', []):
        cat_ids = [str(cid).rstrip('.0') for cid in str(category.get('catId', {}).get(first_letter, '')).split('|')]
        if cat_id_str in cat_ids:
            category_names.append(category.get('category', {}).get('ru', ''))
    return category_names

def update_size_and_weight(products, categories_data):
    categories = categories_data["categories"]
    def get_category_data(category_name):
        for category in categories:
            if category["category"]["ru"] == category_name:
                return {
                    "unitWeight": category.get("unitWeight", 0),
                    "shuttleWeight": category.get("shuttleWeight", 0),
                    "width": category.get("width", 0),
                    "height": category.get("height", 0),
                    "length": category.get("length", 0),
                }
        return None

    for product in products:
        if "categoryName" not in product or "ru" not in product["categoryName"]:
            continue

        category_names = product["categoryName"]["ru"].split("|")
        max_values = {
            "unitWeight": product.get("unitWeight", 0),
            "shuttleWeight": product.get("shuttleWeight", 0),
            "width": product.get("width", 0),
            "height": product.get("height", 0),
            "length": product.get("length", 0),
        }

        for category_name in category_names:
            category_data = get_category_data(category_name)
            if category_data:
                for key in max_values:
                    max_values[key] = max(max_values[key], category_data[key])

        product["unitWeight"] = max_values["unitWeight"]
        product["shuttleWeight"] = max_values["shuttleWeight"]
        product["width"] = max_values["width"]
        product["height"] = max_values["height"]
        product["length"] = max_values["length"]

# Run the function
update_product_categories(products_file, categories_file, excel_file)
